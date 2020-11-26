# -*- coding: utf-8 -*-

# This file is part of the Ingram Micro Cloud Blue Connect connect-cli.
# Copyright (c) 2019-2020 Ingram Micro. All Rights Reserved.

import os
from datetime import datetime
from copy import deepcopy

from click import ClickException
from json import dumps
from urllib import parse

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.colors import Color, WHITE
from openpyxl.worksheet.datavalidation import DataValidation

from tqdm import trange
import requests

from cnctcli.actions.products.constants import (
    ITEMS_COLS_HEADERS,
    PARAMS_COLS_HEADERS,
    MEDIA_COLS_HEADERS,
    CAPABILITIES_COLS_HEADERS,
    STATIC_LINK_HEADERS,
    TEMPLATES_HEADERS,
    PARAM_TYPES,
)
from cnctcli.api.utils import (
    format_http_status,
    handle_http_error,
)
from cnct import ConnectClient, ClientError
from cnct.rql import R


def _setup_cover_sheet(ws, product, location, client):
    ws.title = 'General Information'
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 180
    ws.merge_cells('A1:B1')
    cell = ws['A1']
    cell.fill = PatternFill('solid', start_color=Color('1565C0'))
    cell.font = Font(sz=24, color=WHITE)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.value = 'Product information'
    for i in range(3, 9):
        ws[f'A{i}'].font = Font(sz=14)
        ws[f'B{i}'].font = Font(sz=14, bold=True)
    ws['A3'].value = 'Account ID'
    ws['B3'].value = product['owner']['id']
    ws['A4'].value = 'Account Name'
    ws['B4'].value = product['owner']['name']
    ws['A5'].value = 'Product ID'
    ws['B5'].value = product['id']
    ws['A6'].value = 'Product Name'
    ws['B6'].value = product['name']
    ws['A7'].value = 'Export datetime'
    ws['B7'].value = datetime.now().isoformat()
    ws['A8'].value = 'Product Category'
    ws['B8'].value = product['category']['name']
    ws['A9'].value = 'Product Icon file name'
    ws['B9'].value = product["icon"].split("/")[-1]
    _dump_image(
        f'{location}{product["icon"]}',
        product["id"] + "/" + product["icon"].split("/")[-1]
    )
    ws['A10'].value = 'Product Short Description'
    ws['A10'].alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws['B10'].value = product['short_description']
    ws['B10'].alignment = Alignment(
        wrap_text=True,
    )
    ws['A11'].value = 'Product Detailed Description'
    ws['A11'].alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws['B11'].value = product['detailed_description']
    ws['B11'].alignment = Alignment(
        wrap_text=True,
    )
    ws['A12'].value = 'Buyer short Description'
    ws['B12'].value = product['customer_ui_settings']['description']
    ws['B12'].alignment = Alignment(
        wrap_text=True,
    )
    ws['A13'].value = 'Buyer Getting Started Description'
    ws['B13'].value = product['customer_ui_settings']['getting_started']
    ws['B13'].alignment = Alignment(
        wrap_text=True,
    )

    categories = client.categories.all()
    # Poping categories that does not apply due endpoint has not such filter
    if 'Cloud Services' in categories:
        categories.pop('Cloud Services')
    if 'All Categories' in categories:
        categories.pop('All Categories')

    categories_list = [cat['name'] for cat in categories]
    categories_formula = ','.join(categories_list)
    categories_validation = DataValidation(
        type='list',
        formula1=f'"-,{categories_formula}"',
        allow_blank=False,
    )
    ws.add_data_validation(categories_validation)
    categories_validation.add('B8')


def _dump_image(image_location, image_name):
    image = requests.get(image_location)
    with open(image_name, 'wb') as f:
        f.write(image.content)


def _get_col_limit_by_ws_type(ws_type):
    if ws_type == 'items':
        return 'M'
    elif ws_type == 'params':
        return 'L'
    elif ws_type == 'media':
        return 'F'
    elif ws_type == 'capabilities':
        return 'C'
    elif ws_type == 'static_links':
        return 'D'
    elif ws_type == 'templates':
        return 'F'
    return 'Z'


def _setup_ws_header(ws, ws_type=None):
    if not ws_type:
        ws_type = 'items'

    color = Color('d3d3d3')
    fill = PatternFill('solid', color)
    cels = ws['A1': '{}1'.format(
        _get_col_limit_by_ws_type(ws_type)
    )]
    for cel in cels[0]:
        ws.column_dimensions[cel.column_letter].width = 25
        ws.column_dimensions[cel.column_letter].auto_size = True
        cel.fill = fill
        if ws_type == 'items':
            cel.value = ITEMS_COLS_HEADERS[cel.column_letter]
        elif ws_type == 'params':
            cel.value = PARAMS_COLS_HEADERS[cel.column_letter]
            if cel.value == 'JSON Properties':
                ws.column_dimensions[cel.column_letter].width = 100
        elif ws_type == 'media':
            cel.value = MEDIA_COLS_HEADERS[cel.column_letter]
        elif ws_type == 'capabilities':
            cel.value = CAPABILITIES_COLS_HEADERS[cel.column_letter]
            if cel.value == 'Capability':
                ws.column_dimensions[cel.column_letter].width = 50
        elif ws_type == 'static_links':
            cel.value = STATIC_LINK_HEADERS[cel.column_letter]
            if cel.value == 'Url':
                ws.column_dimensions[cel.column_letter].width = 100
        elif ws_type == 'templates':
            cel.value = TEMPLATES_HEADERS[cel.column_letter]
            if cel.value == 'Content':
                ws.column_dimensions[cel.column_letter].width = 100
            if cel.value == 'Title':
                ws.column_dimensions[cel.column_letter].width = 50


def _calculate_commitment(item):
    period = item.get('period')
    if not period:
        return '-'
    commitment = item.get('commitment')
    if not commitment:
        return '-'
    count = commitment['count']
    if count == 1:
        return '-'

    multiplier = commitment['multiplier']

    if multiplier == 'billing_period':
        if period == 'monthly':
            years = count // 12
            return '{} year{}'.format(
                years,
                's' if years > 1 else '',
            )
        else:
            return '{} years'.format(count)

    # One-time
    return '-'


def _fill_param_row(ws, row_idx, param):
    ws.cell(row_idx, 1, value=param['id']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 2, value=param['name']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 3, value='-').alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 4, value=param['title']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 5, value=param['description']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 6, value=param['phase']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 7, value=param['scope']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 8, value=param['type']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 9, value=param['constraints']['required']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 10, value=param['constraints']['unique']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 11, value=param['constraints']['hidden']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 12, value=_get_json_object_for_param(param)).alignment = Alignment(
        wrap_text=True,
    )


def _get_json_object_for_param(original_param):
    param = deepcopy(original_param)
    del param['id']
    del param['name']
    del param['title']
    del param['description']
    del param['phase']
    del param['scope']
    del param['type']
    del param['constraints']['required']
    del param['constraints']['unique']
    del param['constraints']['hidden']
    del param['position']
    del param['events']

    return dumps(param, indent=4, sort_keys=True)


def _fill_media_row(ws, row_idx, media, location, product):
    ws.cell(row_idx, 1, value=media['position'])
    ws.cell(row_idx, 2, value=media['id'])
    ws.cell(row_idx, 3, value='-')
    ws.cell(row_idx, 4, value=media['type'])
    ws.cell(row_idx, 5, value=media['thumbnail'].split("/")[-1])
    _dump_image(
        f'{location}{media["thumbnail"]}',
        f"./{product}/media/" + media["thumbnail"].split("/")[-1]
    )
    ws.cell(row_idx, 6, value='-' if media['type'] == 'image' else media['url'])


def _fill_template_row(ws, row_idx, template):
    ws.cell(row_idx, 1, value=template['id']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 2, value=template['title']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 3, value='-').alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 4, value=template['scope']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 5, value=template['type']).alignment = Alignment(
        horizontal='left',
        vertical='top',
    )
    ws.cell(row_idx, 6, value=template['body']).alignment = Alignment(
        wrap_text=True,
    )


def _fill_item_row(ws, row_idx, item):
    ws.cell(row_idx, 1, value=item['id'])
    ws.cell(row_idx, 2, value=item['mpn'])
    ws.cell(row_idx, 3, value='-')
    ws.cell(row_idx, 4, value=item['display_name'])
    ws.cell(row_idx, 5, value=item['description'])
    ws.cell(row_idx, 6, value=item['type'])
    ws.cell(row_idx, 7, value=item['precision'])
    ws.cell(row_idx, 8, value=item['unit']['unit'])
    period = item.get('period', 'monthly')
    if period.startswith('years_'):
        period = f"{period.rsplit('_')[-1]} years"
    ws.cell(row_idx, 9, value=period)
    ws.cell(row_idx, 10, value=_calculate_commitment(item))
    ws.cell(row_idx, 11, value=item['status'])
    ws.cell(row_idx, 12, value=item['events']['created']['at'])
    ws.cell(row_idx, 13, value=item['events'].get('updated', {}).get('at'))


def _dump_parameters(ws, client, product_id, param_type, silent):
    _setup_ws_header(ws, 'params')

    rql = R().phase.eq(param_type)

    processed_items = 0
    row_idx = 2

    params = client.products[product_id].parameters.filter(rql)
    count = params.count()

    if count == 0:
        # Product without params is strange, but may exist
        return
    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"{}"'.format(
            ','.join(PARAM_TYPES)
        ),
        allow_blank=False,
    )
    phase_validation = DataValidation(
        type='list',
        formula1='"ordering,fulfillment,configuration"',
        allow_blank=False,
    )
    ordering_fulfillment_scope_validation = DataValidation(
        type='list',
        formula1='"asset,tier1,tier2"',
        allow_blank=False,
    )
    configuration_scope_validation = DataValidation(
        type='list',
        formula1='"product,marketplace,item,item_marketplace"',
        allow_blank=False,
    )
    bool_validation = DataValidation(
        type='list',
        formula1='"True,False"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(phase_validation)
    ws.add_data_validation(ordering_fulfillment_scope_validation)
    ws.add_data_validation(configuration_scope_validation)
    ws.add_data_validation(bool_validation)

    progress = trange(0, count, position=0, disable=silent)

    for param in params:
        progress.set_description(f"Processing {param_type} parameter {param['id']}")
        progress.update(1)
        _fill_param_row(ws, row_idx, param)
        action_validation.add(f'C{row_idx}')
        phase_validation.add(f'F{row_idx}')
        if param['scope'] == 'configuration':
            configuration_scope_validation.add(f'G{row_idx}')
        else:
            ordering_fulfillment_scope_validation.add(f'G{row_idx}')
        type_validation.add(f'H{row_idx}')
        bool_validation.add(f'I{row_idx}')
        bool_validation.add(f'J{row_idx}')
        bool_validation.add(f'K{row_idx}')
        processed_items += 1
        row_idx += 1


def _dump_media(ws, client, product_id, silent, media_location):
    _setup_ws_header(ws, 'media')
    processed_items = 0
    row_idx = 2

    medias = client.products[product_id].media.all()
    count = medias.count()
    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"image,video"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(type_validation)

    progress = trange(0, count, position=0, disable=silent)
    if not os.path.exists(os.getcwd() + f'/{product_id}/media/'):
        os.mkdir(os.getcwd() + f'/{product_id}/media/')
    for media in medias:
        progress.set_description(f"Processing media {media['id']}")
        progress.update(1)
        _fill_media_row(ws, row_idx, media, media_location, product_id)
        action_validation.add(f'C{row_idx}')
        type_validation.add(f'D{row_idx}')
        processed_items += 1
        row_idx += 1


def _dump_external_static_links(ws, product, silent):
    _setup_ws_header(ws, 'static_links')
    row_idx = 2
    count = len(product['customer_ui_settings']['download_links'])
    count = count + len(product['customer_ui_settings']['documents'])

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    link_type = DataValidation(
        type='list',
        formula1='"Download,Documentation"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(link_type)

    progress = trange(0, count, position=0, disable=silent)

    progress.set_description("Processing Static Links")

    for link in product['customer_ui_settings']['download_links']:
        progress.update(1)
        ws.cell(row_idx, 1, value='Download')
        ws.cell(row_idx, 2, value=link['title'])
        ws.cell(row_idx, 3, value='-')
        ws.cell(row_idx, 4, value=link['url'])
        action_validation.add(f'C{row_idx}')
        link_type.add(f'A{row_idx}')
        row_idx += 1

    for link in product['customer_ui_settings']['documents']:
        progress.update(1)
        ws.cell(row_idx, 1, value='Documentation')
        ws.cell(row_idx, 2, value=link['title'])
        ws.cell(row_idx, 3, value='-')
        ws.cell(row_idx, 4, value=link['url'])
        action_validation.add(f'C{row_idx}')
        link_type.add(f'A{row_idx}')
        row_idx += 1


def _dump_capabilities(ws, product, silent):
    _setup_ws_header(ws, 'capabilities')
    progress = trange(0, 1, position=0, disable=silent)
    progress.set_description("Processing Product Capabilities")
    ppu = product['capabilities']['ppu']
    capabilities = product['capabilities']
    tiers = capabilities['tiers']

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    ppu_validation = DataValidation(
        type='list',
        formula1='"Disabled,QT,TR,PR"',
        allow_blank=False,
    )
    disabled_enabled = DataValidation(
        type='list',
        formula1='"Disabled,Enabled"',
        allow_blank=False,
    )
    tier_validation = DataValidation(
        type='list',
        formula1='"Disabled,1,2"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(ppu_validation)
    ws.add_data_validation(disabled_enabled)
    ws.add_data_validation(tier_validation)

    ws['A2'].value = 'Pay-as-you-go support and schema'
    ws['B2'].value = '-'
    ws['C2'].value = (ppu['schema'] if ppu else 'Disabled')
    ppu_validation.add(ws['C2'])
    ws['A3'].value = 'Pay-as-you-go dynamic items support'
    ws['B3'].value = '-'
    ws['C3'].value = (
        ppu['dynamic'] if ppu and 'dynamic' in ppu else 'Disabled'
    )
    disabled_enabled.add(ws['C3'])
    ws['A4'].value = 'Pay-as-you-go future charges support'
    ws['B4'].value = '-'
    ws['C4'].value = (
        ppu['future'] if ppu and 'future' in ppu else 'Disabled'
    )
    disabled_enabled.add(ws['C4'])
    ws['A5'].value = 'Consumption reporting for Reservation Items'
    ws['B5'].value = '-'

    def _get_reporting_consumption(reservation_cap):
        if 'consumption' in reservation_cap and reservation_cap['consumption']:
            return 'Enabled'
        return 'Disabled'

    ws['C5'].value = _get_reporting_consumption(capabilities['reservation'])
    disabled_enabled.add(ws['C5'])
    ws['A6'].value = 'Dynamic Validation of the Draft Requests'
    ws['B6'].value = '-'

    def _get_dynamic_validation_draft(capabilities_cart):
        if 'validation' in capabilities_cart and capabilities['cart']['validation']:
            return 'Enabled'
        return 'Disabled'
    ws['C6'].value = _get_dynamic_validation_draft(capabilities['cart'])
    disabled_enabled.add(ws['C6'])
    ws['A7'].value = 'Dynamic Validation of the Inquiring Form'
    ws['B7'].value = '-'

    def _get_validation_inquiring(capabilities_inquiring):
        if 'validation' in capabilities_inquiring and capabilities_inquiring['validation']:
            return 'Enabled'
        return 'Disabled'

    ws['C7'].value = _get_validation_inquiring(capabilities['inquiring'])
    disabled_enabled.add(ws['C7'])
    ws['A8'].value = 'Reseller Authorization Level'
    ws['B8'].value = '-'
    ws['C8'].value = (
        tiers['configs']['level'] if tiers and 'configs' in tiers else 'Disabled'
    )
    tier_validation.add(ws['C8'])
    ws['A9'].value = 'Tier Accounts Sync'
    ws['B9'].value = '-'
    ws['C9'].value = (
        'Enabled' if tiers and 'updates' in tiers and tiers['updates'] else 'Disabled'
    )
    disabled_enabled.add(ws['C9'])
    ws['A10'].value = 'Administrative Hold'
    ws['B10'].value = '-'
    ws['C10'].value = (
        'Enabled' if 'hold' in capabilities['subscription'] and capabilities['subscription']['hold'] else 'Disabled'
    )
    disabled_enabled.add(ws['C10'])
    idx = 2
    while idx < 11:
        action_validation.add(f'B{idx}')
        idx = idx + 1
    progress.update(1)


def _dump_templates(ws, client, product_id, silent):
    _setup_ws_header(ws, 'templates')

    processed_items = 0
    row_idx = 2

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    scope_validation = DataValidation(
        type='list',
        formula1='"asset,tier1,tier2"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"fulfillment,inquire"',
        allow_blank=False,
    )
    ws.add_data_validation(action_validation)
    ws.add_data_validation(scope_validation)
    ws.add_data_validation(type_validation)

    templates = client.products[product_id].templates.all()
    count = templates.count()

    progress = trange(0, count, position=0, disable=silent)

    for template in templates:
        progress.set_description(f"Processing Template {template['id']}")
        progress.update(1)
        if 'type' in template:
            _fill_template_row(ws, row_idx, template)
            action_validation.add(f'C{row_idx}')
            scope_validation.add(f'D{row_idx}')
            type_validation.add(f'E{row_idx}')
            row_idx += 1

        processed_items += 1


def _dump_items(ws, client, product_id, silent):
    _setup_ws_header(ws, 'items')

    processed_items = 0
    row_idx = 2

    items = client.products[product_id].items.all()
    count = items.count()

    if count == 0:
        raise ClickException(f"The product {product_id} doesn't have items.")

    action_validation = DataValidation(
        type='list',
        formula1='"-,create,update,delete"',
        allow_blank=False,
    )
    type_validation = DataValidation(
        type='list',
        formula1='"reservation,ppu"',
        allow_blank=False,
    )
    period_validation = DataValidation(
        type='list',
        formula1='"onetime,monthly,yearly,2 years,3 years,4 years,5 years"',
        allow_blank=False,
    )

    precision_validation = DataValidation(
        type='list',
        formula1='"integer,decimal(1),decimal(2),decimal(4),decimal(8)"',
        allow_blank=False,
    )

    commitment_validation = DataValidation(
        type='list',
        formula1='"-,1 year,2 years,3 years,4 years,5 years"',
        allow_blank=False,
    )

    ws.add_data_validation(action_validation)
    ws.add_data_validation(type_validation)
    ws.add_data_validation(period_validation)
    ws.add_data_validation(precision_validation)
    ws.add_data_validation(commitment_validation)

    progress = trange(0, count, position=0, disable=silent)

    for item in items:
        progress.set_description(f"Processing item {item['id']}")
        progress.update(1)
        _fill_item_row(ws, row_idx, item)
        action_validation.add(f'C{row_idx}')
        type_validation.add(f'F{row_idx}')
        precision_validation.add(f'G{row_idx}')
        period_validation.add(f'I{row_idx}')
        commitment_validation.add(f'J{row_idx}')
        processed_items += 1
        row_idx += 1


def dump_product(api_url, api_key, product_id, output_file, silent):
    if not os.path.exists(os.getcwd() + f'/{product_id}'):
        os.mkdir(os.getcwd() + f'/{product_id}')
    elif not os.path.isdir(os.getcwd() + f'/{product_id}'):
        raise ClickException(
            "Exists a file with product name but a directory is expected, please rename it"
        )
    if not output_file:
        output_file = os.path.abspath(
            os.path.join(f'./{product_id}/', f'{product_id}.xlsx'),
        )
    try:
        client = ConnectClient(api_key=api_key, endpoint=api_url)
        product = client.products[product_id].get()
        wb = Workbook()
        connect_api_location = parse.urlparse(api_url)
        media_location = connect_api_location.scheme + "://" + connect_api_location.netloc
        _setup_cover_sheet(
            wb.active,
            product,
            media_location,
            client,
        )

        _dump_capabilities(wb.create_sheet('Capabilities'), product, silent)
        _dump_external_static_links(wb.create_sheet('Buyer Static Resources'), product, silent)
        _dump_media(wb.create_sheet('Media'), client, product_id, silent, media_location)
        _dump_templates(wb.create_sheet('Templates'), client, product_id, silent)
        _dump_items(wb.create_sheet('Items'), client, product_id, silent)
        _dump_parameters(
            wb.create_sheet('Params Ordering'),
            client,
            product_id,
            'ordering',
            silent
        )
        _dump_parameters(
            wb.create_sheet('Params Fulfillment'),
            client,
            product_id,
            'fulfillment',
            silent
        )
        _dump_parameters(
            wb.create_sheet('Params Configuration'),
            client,
            product_id,
            'configuration',
            silent
        )
        wb.save(output_file)

    except ClientError as error:
        status = format_http_status(error.status_code)
        if error.status_code == 404:
            raise ClickException(f'{status}: Product {product_id} not found.')

        handle_http_error(error)

    return output_file
